#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import a filled 项目配置填写表.xlsx into a profiles/<slug>.yaml scaffold.

Deterministic mechanical pass only (lookup / mapping / split / baseline / flag gaps):
  - auto-detect which column holds the answers: D (correct, yellow column) OR
    C (the 燕云 example column the filler accidentally overwrote)
  - map Q1..Q12 -> yaml fields
  - inject the 古风/仙侠 shared baseline (surname extractor / address suffixes /
    generic denoise) when the genre matches
  - mark every empty or judgment-needed field with TODO(PM) / TODO(AI)

Semantic polish is intentionally NOT done here and is left to an AI/human pass:
  - normalizing Q7 into short single-pick category labels
  - turning Q4/Q6 into clean core_principle prose + project-specific name rules
  - parsing Q10 positive examples into structured few-shot pairs
The script never fabricates game lore: unknown project-specific fields stay TODO.

Pairs with scripts/gen_project_form.py (which generates the blank form).

Usage:
  python scripts/import_project_form.py <filled_form.xlsx> --slug gujian
  python scripts/import_project_form.py <form.xlsx> --slug gujian --out /tmp/x.yaml --force
"""
import argparse
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = ROOT / "docs" / "项目配置填写表.xlsx"   # blank form; its col C = example sentinels
SHEET = "项目信息填写表"

# Genres that share the 古风/仙侠 Chinese-name machinery below.
GUWU_KEYWORDS = ["古风", "武侠", "仙侠", "古代", "江湖", "国风", "古典", "东方"]

# --- 古风/仙侠 shared baseline (used to fill genre-generic fields left blank) ---
SURNAMES = (
    "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳"
    "酆鲍史唐费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮下齐康伍余元卜顾孟平黄和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊"
    "纪舒屈项祝董梁杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林刁钟徐邱骆高夏蔡田樊胡凌霍虞万支柯昝管卢莫经房裘缪干解应宗丁宣贲邓郁单"
    "杭洪包诸左石崔吉钮龚程嵇邢滑裴陆荣翁荀羊於惠甄麴家封芮羿储靳汲邴糜松井段富巫乌焦巴弓牧隗山谷车侯宓蓬全郗班仰秋仲伊宫宁仇栾"
    "暴甘钭厉戎祖武符刘景詹束龙叶幸司韶郜黎蓟薄印宿白怀蒲邰从鄂索咸籍赖卓蔺屠蒙池乔阴鬱胥能苍双闻莘党翟谭贡劳逄姬申扶堵冉宰郦雍"
    "卻璩桑桂濮牛寿通边扈燕冀郏浦尚农温别庄晏柴瞿阎充慕连茹习宦艾鱼容向古易慎戈廖庾终暨居衡步都耿满弘匡国文寇广禄阙东欧殳沃利蔚"
    "越夔隆师巩厍聂晁勾敖融冷訾辛阚那简饶空曾毋沙乜养鞠须丰巢关蒯相查後荆红游竺权逮盍益桓公"
)
BASELINE = {
    "address_suffixes": ["大哥", "大嫂", "大婶", "大娘", "大爷", "大姐", "大伯", "师妹", "师弟",
                         "师姐", "师兄", "师叔", "师伯", "嫂子", "姑娘", "娘子", "前辈", "兄", "哥"],
    "filterable_categories": ["通用物品", "通用NPC"],
    "extraction_notes": [
        "中文人名模式强制提取：老X、小X、阿X、X嫂、X婶、X伯、X叔、X爷、X娘、X姑、X儿、X子、X哥、X弟——一个不漏",
        "常规通称不提取：先生、大侠、少侠、公子、姑娘、弟子、同门、前辈、掌门——除非是专有名字的一部分",
        "通用场所不提取：书房、医馆、客栈、夜市等无专有名称的场所",
        "通用原材料不提取：柴火、木炭、绳子等日常材料",
        "节气/泛时间词不提取：子时、春分等",
    ],
    "exclude": [
        "通用日常物品（如：木炭、草鞋、火把、柴火、绳子）",
        "通用称谓/泛称（如：大侠、公子、弟子、长老、前辈——非专有名字的一部分时）",
        "成语/固定短语",
        "时间词（如：子时、昨夜）",
        "无专有名称的通用场所（如：书房、医馆、客栈）",
    ],
    "negatives": ["用木炭生火，草鞋踩在石板上", "长老说弟子们不得擅自出门，前辈也不例外", "他一飞冲天，大鹏展翅般冲出重围"],
    "translation_rules": [
        "角色名/NPC名 → 音译（拼音）",
        "技能/招式名 → 意译，传达含义",
        "地点/建筑名 → 意译为主，必要时音译+意译结合",
        "物品/道具名 → 简洁意译",
        "头衔称谓 → 意译（如公子→Master, 长老→Elder）",
        "每个术语一行，格式：中文术语 → 英文译文",
    ],
}
GENERIC_PRINCIPLE = (
    "文本中出现的中文人名 / NPC 名——无论主线、支线、路人、单次提及——一律提取。\n"
    "人物重要性、出场次数、是否主线，都不影响：出现即术语。\n"
)


# ----------------------------- form reading -----------------------------

def open_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb[SHEET] if SHEET in wb.sheetnames else wb.active


def index_questions(ws):
    """Map question number -> row, by matching 'N.' in column A. Robust to
    shifted rows and to a damaged section-header cell."""
    qrow = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        m = re.match(r"\s*(\d+)\s*[\.、]", str(a))
        if m:
            qrow[int(m.group(1))] = r
    return qrow


def _clean(v):
    return str(v).strip() if v not in (None, "") else ""


def build_example_sentinels():
    """Col C of the blank template = the 燕云 example per question. Used to tell
    'C still holds the example' (unfilled) from 'filler overwrote C' (answer)."""
    if not TEMPLATE.exists():
        return {}
    ws = open_sheet(TEMPLATE)
    qrow = index_questions(ws)
    return {q: _clean(ws.cell(r, 3).value) for q, r in qrow.items()}


def extract_answers(ws, qrow, sentinels):
    """For each question return (answer, source). Prefer D; else C if it differs
    from the example sentinel (the overwritten-column case)."""
    ans, src = {}, {}
    for q, r in qrow.items():
        c = _clean(ws.cell(r, 3).value)
        d = _clean(ws.cell(r, 4).value)
        if d:
            ans[q], src[q] = d, "D"
        elif c and c != sentinels.get(q, "\0"):
            ans[q], src[q] = c, "C"
        else:
            ans[q], src[q] = "", "空"
    return ans, src


# ----------------------------- text helpers -----------------------------

def split_lines(text):
    out = []
    for ln in str(text).splitlines():
        ln = ln.strip().lstrip("·-•*◦・　 \t").strip()
        if ln:
            out.append(ln)
    return out


def strip_paren(s):
    return re.sub(r"[（(][^）)]*[）)]", "", s).strip()


TERM_RE = re.compile(r"([^\s、，,（()「」『』]+)\s*[（(]\s*([^）)]+?)\s*[）)]")
ARROW_RE = re.compile(r"\s*(?:→|->|=>|⇒)\s*")


def parse_positive(lines):
    """Best-effort: '「原文」→ 术语(分类)、术语(分类)' -> few-shot pair.
    Unparseable lines are returned as leftovers for a TODO dump."""
    pairs, leftover = [], []
    for ln in lines:
        parts = ARROW_RE.split(ln, maxsplit=1)
        if len(parts) != 2:
            leftover.append(ln)
            continue
        src = parts[0].strip().strip("「」『』\"' ")
        terms = [(t.strip(), cat.strip()) for t, cat in TERM_RE.findall(parts[1])]
        if src and terms:
            pairs.append((src, terms))
        else:
            leftover.append(ln)
    return pairs, leftover


def parse_negative(lines):
    """Each line -> an empty-output few-shot. Strips quotes / trailing '→ 无术语'."""
    negs = []
    for ln in lines:
        s = ARROW_RE.split(ln, maxsplit=1)[0].strip().strip("「」『』\"' ")
        if s:
            negs.append(s)
    return negs


# ----------------------------- yaml emit -----------------------------

def yq(s):
    return '"' + str(s).replace("\\", "\\\\").replace('"', '\\"') + '"'


def emit_list(key, items, marker):
    tag = f"   # {marker}" if marker else ""
    if not items:
        return f"{key}: []{tag}\n"
    s = f"{key}:{tag}\n"
    for it in items:
        s += f"  - {yq(it)}\n"
    return s


def emit_block(key, text, marker):
    tag = f"   # {marker}" if marker else ""
    s = f"{key}: |{tag}\n"
    for ln in str(text).rstrip("\n").splitlines():
        s += f"  {ln}\n"
    return s


def emit_fewshot(positives, negatives, marker):
    s = f"fewshot_examples:   # {marker}\n"
    for src, terms in positives:
        s += f"  - input: {yq(src)}\n    output:\n"
        for t, cat in terms:
            s += f"      - term: {yq(t)}\n        category: {yq(cat)}\n"
    for src in negatives:
        s += f"  - input: {yq(src)}\n    output: []\n"
    return s


def build_yaml(slug, ans, src, guwu):
    """Assemble the commented profile scaffold text."""
    def used(q):
        tag = "PM" if ans.get(q) else "默认"
        where = {"C": " ·C列(填错)", "D": " ·D列"}.get(src.get(q, ""), "")
        return f"[{tag}] Q{q}{where}"

    game_name = ans.get(1, "") or slug
    style = ans.get(2, "")
    game_type = f"{style}（{game_name}）" if style and game_name else (style or game_name)
    task = ans.get(3, "") or f"{game_name}游戏术语抽取任务"

    include = split_lines(ans.get(4, ""))
    exclude = split_lines(ans.get(5, ""))
    notes = split_lines(ans.get(6, ""))
    cats_raw = split_lines(ans.get(7, ""))
    cats = [strip_paren(x).lstrip("所有").lstrip("全部").strip() or x for x in cats_raw]
    trules = split_lines(ans.get(9, ""))
    pos, pos_leftover = parse_positive(split_lines(ans.get(10, "")))
    negs = parse_negative(split_lines(ans.get(11, "")))
    misc = ans.get(12, "")

    L = []
    c_qs = [q for q in sorted(src) if src[q] == "C"]
    L.append(f"# {game_name} 术语抽取 profile — 由 scripts/import_project_form.py 生成")
    L.append("#   [PM]=填写表真实答案  [默认]=古风/仙侠通用基线待校准  [TODO]=空缺/需判断，勿照搬基线编造")
    if c_qs:
        L.append(f"#   ⚠ 填写表把 Q{c_qs} 的答案填进了 C 列(示例列)而非 D 列；已按 C 列读取")
    L.append("")

    gt_tag = "PM" if (ans.get(1) or ans.get(2)) else "默认"
    L.append(f"game_type: {yq(game_type)}      # [{gt_tag}] Q1+Q2")
    L.append(f"task_description: {yq(task)}      # {used(3)}")
    L.append("")

    if include:
        L.append(emit_block("core_principle", GENERIC_PRINCIPLE,
                            "[默认] 通用强口径；古剑专属人名构词/单字名/排行命名规则待 Q6 回填后细化").rstrip("\n"))
    else:
        L.append('core_principle: ""   # [TODO] Q4 空，无法定口径')
    L.append("")

    if guwu:
        L.append("rule_extractors:   # [默认] 百家姓确定性人名提取器，通用中文姓名")
        L.append("  surname_names:")
        L.append(f"    surnames: {yq(SURNAMES)}")
        L.append("    min_len: 2")
        L.append("    max_len: 3")
        L.append("")
        L.append(emit_list("address_suffixes", BASELINE["address_suffixes"], "[默认] 通用称谓后缀").rstrip("\n"))
        L.append("")
        L.append(emit_list("filterable_categories", BASELINE["filterable_categories"], "[默认] 命中即整体过滤").rstrip("\n"))
        L.append("")

    if notes:
        L.append(emit_list("extraction_notes", notes, used(6)).rstrip("\n"))
    elif guwu:
        L.append(emit_list("extraction_notes", BASELINE["extraction_notes"],
                          "[默认] Q6 空，用通用降噪；回填后追加古剑专属硬规则").rstrip("\n"))
    else:
        L.append("extraction_notes: []   # [TODO] Q6 空")
    L.append("")

    L.append("extract_examples:")
    if include:
        L.append("  " + emit_list("include", include, used(4)).rstrip("\n").replace("\n", "\n  "))
    else:
        L.append("  include: []   # [TODO] Q4 空")
    exc = exclude if exclude else (BASELINE["exclude"] if guwu else [])
    exc_mark = used(5) if exclude else ("[默认] Q5 空，通用噪音待校准" if guwu else "[TODO] Q5 空")
    L.append("  " + emit_list("exclude", exc, exc_mark).rstrip("\n").replace("\n", "\n  "))
    L.append("")

    neg_items = negs if negs else (BASELINE["negatives"] if guwu else [])
    fs_mark = (used(10) + "/" + used(11)) if (pos or negs) else "[默认/TODO] Q10 正例缺(掉召回)；负例为通用占位"
    L.append(emit_fewshot(pos, neg_items, fs_mark).rstrip("\n"))
    if pos_leftover:
        L.append("# [TODO(AI)] Q10 未能解析的正例原文，请手工补成 input/term/category：")
        for ln in pos_leftover:
            L.append(f"#   {ln}")
    L.append("")

    if cats:
        cats_full = cats + (["通用NPC", "通用物品"] if guwu else [])
        L.append(emit_list("term_categories", cats_full,
                          used(7) + " — [TODO(AI)] 脚本粗提取，请规整为简短可单选标签").rstrip("\n"))
    else:
        L.append("term_categories: []   # [TODO] Q7 空")
    L.append("")

    tr = trules if trules else (BASELINE["translation_rules"] if guwu else [])
    tr_mark = used(9) if trules else ("[默认] Q8默认中→英；Q9空，通用仙侠译法待校准" if guwu else "[TODO] Q9 空")
    L.append(emit_list("translation_rules", tr, tr_mark).rstrip("\n"))

    if misc and misc != "（如：某些章节用方言；客户要求保留繁体专名……）":
        L.append("")
        L.append(f"# 其他说明 (Q12)：{misc}")

    return "\n".join(L) + "\n"


# ----------------------------- main -----------------------------

def main():
    ap = argparse.ArgumentParser(description="Import a filled project form into a profile scaffold.")
    ap.add_argument("form", help="path to the filled 项目配置填写表.xlsx")
    ap.add_argument("--slug", required=True, help="profile slug, e.g. gujian -> profiles/gujian.yaml")
    ap.add_argument("--out", help="output path (default profiles/<slug>.yaml)")
    ap.add_argument("--force", action="store_true", help="overwrite if output exists")
    args = ap.parse_args()

    form = Path(args.form)
    if not form.exists():
        sys.exit(f"ERROR: form not found: {form}")
    out = Path(args.out) if args.out else ROOT / "profiles" / f"{args.slug}.yaml"
    if out.exists() and not args.force:
        sys.exit(f"ERROR: {out} exists. Use --force to overwrite.")

    ws = open_sheet(form)
    qrow = index_questions(ws)
    if not qrow:
        sys.exit("ERROR: no 'N.' question rows found — is this the right form?")
    sentinels = build_example_sentinels()
    ans, src = extract_answers(ws, qrow, sentinels)
    style = ans.get(2, "")
    guwu = any(k in style for k in GUWU_KEYWORDS)

    text = build_yaml(args.slug, ans, src, guwu)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(text, encoding="utf-8")

    filled = [q for q in sorted(ans) if ans[q]]
    empty = [q for q in sorted(ans) if not ans[q]]
    cols = sorted({src[q] for q in filled})
    print(f"OK -> {out}")
    print(f"  answer column(s): {', '.join(cols) or 'none'}")
    print(f"  filled Q: {filled}")
    print(f"  empty  Q: {empty}  (left as 默认/TODO)")
    print(f"  genre baseline (古风/仙侠): {'injected' if guwu else 'NOT injected'}  [style={style!r}]")
    print("  next: AI/human polish term_categories labels, core_principle, Q10 positives.")


if __name__ == "__main__":
    main()
