#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the Gujian (古剑奇谭四 / Project SOL) glossary xlsx — the pipeline's
★命门 term-table input (load_glossary reads col0=中文, col1=英文; row1=header).

Sources merged:
  [术语表]  the official term table the PM provided (CN/EN/category/importance)
  [世界观]  CN→EN pairs extracted from the bilingual "Project SOL info" worldview doc
            (official translations embedded in the doc; flagged for review)

Slash variants ("古剑/剑") are expanded to one row each, sharing EN — so the
matcher hits every surface form. Output columns: 中文术语 | 英文翻译 | 分类 |
重要度 | 来源 | 备注. Pipeline reads only the first two; the rest are for humans.

Usage: python scripts/build_gujian_glossary.py [--out test_file/gujian_glossary.xlsx]
"""
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent

# (中文, 英文, 分类, 重要度, 来源)  — slash in 中文 => expanded to multiple rows
TERMS = [
    # ── 角色 (Characters) ──
    ("晏闻", "Yan Wen / Ash", "角色名", "★★★", "术语表"),
    ("姜玄", "Jiang Xuan", "角色名", "★★★", "术语表"),
    ("楚鸣", "Chu Ming", "角色名", "★★★", "术语表"),
    ("由庚", "You Geng", "角色名", "★★★", "术语表"),       # 剑灵 sword spirit
    ("魁苓", "Kui Ling", "角色名", "★★", "术语表"),
    ("楚长生", "Chu Changsheng", "角色名", "", "世界观"),
    ("襄垣", "Xiangyuan", "角色名", "", "世界观"),
    ("祝方", "Zhu Fang", "角色名", "", "世界观"),           # 姜玄化名 alias
    ("尚先生", "Mr. Shang", "角色名", "", "世界观"),
    ("蒙鸠", "Meng Jiu", "角色名", "", "世界观"),
    ("蒙恬", "Meng Tian", "角色名", "", "世界观"),
    ("胡亥", "Hu Hai", "角色名", "", "世界观"),
    ("秦王政/秦始皇", "Emperor of Qin", "角色名", "", "世界观"),
    ("轩辕氏", "Xuanyuan", "角色名", "", "世界观"),
    ("公孙", "Gongsun", "角色名", "", "世界观"),
    ("力牧", "Limu", "角色名", "", "世界观"),
    ("风后", "Fenghou", "角色名", "", "世界观"),
    # ── 神祇 (Gods) ──
    ("伏羲", "Fu Xi", "神祇", "★", "术语表"),
    ("女娲", "Nv Wa / Nuwa", "神祇", "★★", "术语表"),
    ("阎罗", "Yan Luo / Yama", "神祇", "★★", "术语表"),
    ("烛龙", "Aurogon", "神祇", "★★", "术语表"),
    ("商羊", "Shang Yang", "神祇", "★★", "术语表"),
    ("神农", "Shennong", "神祇", "", "世界观"),
    ("蓐收", "Rushou", "神祇", "", "世界观"),
    ("句芒", "Goumang", "神祇", "", "世界观"),
    ("共工", "Gonggong", "神祇", "", "世界观"),
    ("祝融", "Zhurong", "神祇", "", "世界观"),
    ("后土", "Houtu", "神祇", "", "世界观"),
    ("飞廉", "Feilian", "神祇", "", "世界观"),
    ("絜鉤", "Xiegou", "神祇", "", "世界观"),
    ("盘古", "Pangu / the Great Giant", "神祇", "", "世界观"),
    ("祖江", "Zujiang / the Great Serpent", "神祇", "", "世界观"),
    ("钟鼓", "Zhonggu", "神祇", "", "世界观"),
    ("太白星君", "Taibai Star Lord", "神祇", "", "世界观"),
    ("太昊", "Taihao", "神祇", "", "世界观"),
    ("炎帝", "Yandi", "神祇", "", "世界观"),
    ("少昊", "Shaohao", "神祇", "", "世界观"),
    ("颛顼", "Zhuanxu", "神祇", "", "世界观"),
    # ── 神兽 / 魔族 / 种族 ──
    ("朱雀", "the Vermillion Bird", "神兽", "★★★", "术语表"),
    ("守护神兽", "the Guardian Beast", "神兽", "", "世界观"),
    ("蚩尤", "Chi You", "魔族", "★★★", "术语表"),
    ("朔", "Shuo", "魔族", "★★★", "术语表"),
    ("大天魔", "the greater demon", "魔族", "", "世界观"),
    ("魔族", "the Demons", "种族", "★★", "术语表"),
    ("神族", "the Gods", "种族", "★★", "术语表"),
    ("人族", "the Humans", "种族", "★★", "术语表"),
    ("妖族", "the Monsters", "种族", "★★", "术语表"),
    # ── 部族 / 势力 ──
    ("九黎部族", "the Jiuli Clan", "部族势力", "★★★", "术语表"),
    ("安邑部族", "the Anyi Tribe", "部族势力", "", "世界观"),
    ("有熊部落", "the Youxiong tribe", "部族势力", "", "世界观"),
    ("守陵人", "the tomb keepers", "部族势力", "", "世界观"),
    ("巫臷民", "the Wulei people", "部族势力", "", "世界观"),
    # ── 地名 / 秘境 ──
    ("菽山", "Mount Shu", "地名秘境", "★★★", "术语表"),
    ("忘川", "Wang Chuan River / River of Oblivion", "地名秘境", "★★★", "术语表"),
    ("神陵", "the Divine Mausoleum", "地名秘境", "", "世界观"),
    ("菽山神陵", "the Mausoleum of Mount Shu", "地名秘境", "", "世界观"),
    ("不周山", "Mount Buzhou", "地名秘境", "", "世界观"),
    ("洪涯境", "Hongya", "地名秘境", "", "世界观"),
    ("安邑", "Anyi", "地名秘境", "", "世界观"),
    ("龙渊", "Longyuan", "地名秘境", "", "世界观"),
    ("青邱", "Qingqiu", "地名秘境", "", "世界观"),
    ("堙山", "Yin Mountain", "地名秘境", "", "世界观"),
    ("都广/黑水都广", "Duguang", "地名秘境", "", "世界观"),
    ("中原", "the Central Plains", "地名秘境", "", "世界观"),
    ("中土", "the Middle Earth", "地名秘境", "", "世界观"),
    ("涿鹿", "Zhuolu", "地名秘境", "", "世界观"),
    ("咸阳", "Xianyang", "地名秘境", "", "世界观"),
    ("洞天", "the cave", "地名秘境", "", "世界观"),
    ("地府", "the Underworld Palace", "地名秘境", "", "世界观"),
    ("轮回之井/轮回井", "the Well of Reincarnation", "冥界轮回", "", "世界观"),
    ("沙漠长城", "the Great Wall of the desert", "地名秘境", "", "世界观"),
    ("天柱", "the Sky Pillar", "地名秘境", "", "世界观"),
    # ── 三界 / 世界概念 ──
    ("常世", "the Essential World", "灵力概念", "★", "术语表"),
    ("魔界/魔域", "the Outer World", "灵力概念", "★★", "术语表"),
    ("天界", "the Upper Realm", "灵力概念", "★", "术语表"),
    ("人界", "the Middle Realm", "灵力概念", "★", "术语表"),
    ("地界", "the Under Realm", "灵力概念", "★★", "术语表"),
    ("三界", "the Three Realms", "灵力概念", "", "世界观"),
    # ── 灵力 / 概念 ──
    ("地脉", "Earthveins", "灵力概念", "★★★", "术语表"),
    ("清气", "Clear Ki", "灵力概念", "★★", "术语表"),
    ("浊气", "Turbid Ki", "灵力概念", "★★", "术语表"),
    ("混沌", "chaos", "灵力概念", "", "世界观"),
    ("阴阳五行", "Yin-Yang and the Five Elements", "灵力概念", "", "世界观"),
    ("八卦", "the Eight Diagrams", "灵力概念", "", "世界观"),
    ("三魂七魄", "three souls and seven spirits", "灵力概念", "", "世界观"),
    # ── 法器 / 神器 ──
    ("古剑/剑", "SWORD", "法器神器", "★★★", "术语表"),
    ("箍命索", "Bandages of Sealing", "法器神器", "★★★", "术语表"),
    ("镇石", "Warding Stone", "法器神器", "★★★", "术语表"),
    ("寰印未央", "Seal of Endless", "法器神器", "★★★", "术语表"),
    ("火羽", "the fire feather", "法器神器", "", "世界观"),
    ("鸟形玉佩/玉佩", "the jade pendant", "法器神器", "", "世界观"),
    ("魂器", "soul vessel", "法器神器", "", "世界观"),
    ("创世之火", "the Prime Flame", "法器神器", "", "世界观"),
    ("建木", "Jianmu", "法器神器", "", "世界观"),
    # ── 冥界 / 轮回 ──
    ("地界司判", "Judge of the Under Realm", "称号官职", "★★★", "术语表"),
    ("马面", "Horse", "冥界轮回", "★★★", "术语表"),
    ("牛头", "Ox", "冥界轮回", "★★★", "术语表"),
    ("鬼差", "ghost messengers", "冥界轮回", "", "世界观"),
    ("怨灵", "resentful spirit", "冥界轮回", "", "世界观"),
    ("荒魂", "barren soul", "冥界轮回", "", "世界观"),
    # ── 术法 / 巫术 ──
    ("楔魂之术", "the Soul Embedding Spell", "术法", "", "世界观"),
    ("命魂牵引之术/牵引命魂", "the Spell of Soul-guiding", "术法", "", "世界观"),
    ("奇门遁甲", "Qimen Dunjia", "术法", "", "世界观"),
    # ── 称号 / 官职 ──
    ("大巫祝", "Arch Shaman", "称号官职", "★★", "术语表"),
    ("魔帝", "the Demon Emperor", "称号官职", "", "世界观"),
    ("黄帝/轩辕黄帝", "the Yellow Emperor", "称号官职", "", "世界观"),
    ("三皇", "the Three Emperors", "称号官职", "", "世界观"),
    ("金神", "the god of metal", "称号官职", "", "世界观"),
    ("木神", "the god of wood", "称号官职", "", "世界观"),
    ("水神", "the god of water", "称号官职", "", "世界观"),
    ("火神", "the god of fire", "称号官职", "", "世界观"),
    ("土神", "the god of land", "称号官职", "", "世界观"),
    ("风神", "the god of wind", "称号官职", "", "世界观"),
    ("雨神", "the god of rain", "称号官职", "", "世界观"),
    # ── 军事 / 秦 ──
    ("秦俑军团", "the Qin Terracotta Warriors", "部族势力", "", "世界观"),
    ("俑卫", "the terracotta warriors", "部族势力", "", "世界观"),
    ("飞天舰队", "the Flying Warships", "部族势力", "", "世界观"),
    # ── 历史 / 时代 / 事件 ──
    ("战国", "the Warring States", "玩法机制", "", "世界观"),
    ("涿鹿之战", "the Battle of Zhuolu", "玩法机制", "", "世界观"),
    ("九黎之难", "the Fall of Jiuli", "玩法机制", "", "世界观"),
]

HEADER = ["中文术语", "英文翻译", "分类", "重要度", "来源", "备注"]


def expand(rows):
    out = []
    seen = set()
    for cn, en, cat, imp, src in rows:
        for one in cn.split("/"):
            one = one.strip()
            if one and one not in seen:
                seen.add(one)
                out.append([one, en, cat, imp, src, ""])
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(ROOT / "test_file" / "gujian_glossary.xlsx"))
    args = ap.parse_args()

    rows = expand(TERMS)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "glossary"
    ws.append(HEADER)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(r)
    widths = [16, 34, 12, 9, 8, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)

    from collections import Counter
    cats = Counter(r[2] for r in rows)
    src = Counter(r[4] for r in rows)
    print(f"OK -> {args.out}")
    print(f"  rows: {len(rows)} terms (header row excluded from glossary)")
    print(f"  by source: {dict(src)}")
    print(f"  by category: {dict(cats)}")
    print("  pipeline uses col0=中文术语, col1=英文翻译; 世界观-sourced rows = review-then-trust")


if __name__ == "__main__":
    main()
